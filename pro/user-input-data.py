import webbrowser
import sys
import os
import time
import requests
import subprocess
from pathlib import Path

from playwright.sync_api import sync_playwright

from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QCheckBox,
    QPushButton, QVBoxLayout, QHBoxLayout, QFrame, QMessageBox,
    QScrollArea, QInputDialog, QTextEdit, QTabWidget
)
from PyQt6.QtCore import Qt, QPoint, QTimer, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QPainter, QPolygon, QColor, QCursor, QIcon


# ── Base path ─────────────────────────────────────────────────────
def _app_base():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

# ── SerpAPI key loader ────────────────────────────────────────────
def get_local_api_key():
    key_file_path = os.path.join(_app_base(), "serp_api.txt")
    if os.path.exists(key_file_path):
        try:
            with open(key_file_path, "r", encoding="utf-8") as f:
                return f.read().strip()
        except: pass
    return os.environ.get("SERPAPI_KEY")

def fetch_live_serp_credits():
    api_key = get_local_api_key()
    if not api_key:
        return "Missing Key"
    try:
        r = requests.get("https://serpapi.com/account.json",
                         params={"api_key": api_key}, timeout=3)
        if r.status_code == 200:
            return str(r.json().get("plan_searches_left", 0))
    except: pass
    return "Offline"

# ── Anthropic key helper ──────────────────────────────────────────
def get_anthropic_key_status():
    """Returns (found: bool, masked_key: str)"""
    key_file = os.path.join(_app_base(), "anthropic_api.txt")
    if os.path.exists(key_file):
        try:
            key = open(key_file, encoding="utf-8").read().strip()
            if key:
                return True, key[:8] + "••••••••"
        except: pass
    return False, ""

# ── Passkey cache helpers ─────────────────────────────────────────
def _auth_file_path():
    return os.path.join(_app_base(), ".tvauth")

def _load_saved_passkey():
    try:
        with open(_auth_file_path(), 'r', encoding='utf-8') as f:
            return f.read().strip()
    except: return None

def _save_passkey(passkey):
    try:
        with open(_auth_file_path(), 'w', encoding='utf-8') as f:
            f.write(passkey)
    except: pass

def _clear_saved_passkey():
    try: os.remove(_auth_file_path())
    except: pass

def _validate_passkey_cloud(passkey_clean):
    CSV_URL = (
        "https://docs.google.com/spreadsheets/d/"
        "1_mHFrZcnhupYNU2FA9B1I5DEFNerNsIwW61lX_ygPHs/export?format=csv&gid=0"
    )
    try:
        resp = requests.get(CSV_URL, timeout=4)
        if resp.status_code == 200:
            active_keys = set()
            for line in resp.text.splitlines():
                for seg in line.split(','):
                    t = seg.strip().replace('"','').replace("'",'')
                    if t: active_keys.add(t.lower())
            return passkey_clean.lower() in active_keys
        return None
    except: return None


# ── Logo widget ───────────────────────────────────────────────────
class LogoWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(50, 50)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setBrush(QColor("#6B1D66"))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawRect(0, 0, self.width(), self.height())
        painter.setBrush(QColor("#E397E1"))
        pts = QPolygon([QPoint(0,0), QPoint(self.width(),0),
                        QPoint(self.width()//2, int(self.height()*0.46))])
        painter.drawPolygon(pts)


# ── Worker: search scraper ────────────────────────────────────────
class ScraperWorker(QThread):
    finished_signal = pyqtSignal(dict)
    error_signal    = pyqtSignal(str)
    progress_signal = pyqtSignal(str)

    def __init__(self, search_query, allowed_ratings, target_city):
        super().__init__()
        self.search_query    = search_query
        self.allowed_ratings = allowed_ratings
        self.target_city     = target_city

    def run(self):
        try:
            URL = (f"https://raw.githubusercontent.com/skengeneral/TimeVehicle-basic"
                   f"/main/scraper_engine.py?t={int(time.time())}")
            self.progress_signal.emit("🌐 Fetching latest engine from cloud...")
            response = requests.get(URL, headers={'User-Agent':'Mozilla/5.0'}, timeout=15)
            if response.status_code == 200:
                patched = "from pathlib import Path\n" + response.text.replace(
                    "current_dir = Path(os.path.dirname(os.path.abspath(__file__)))",
                    "current_dir = Path(os.getcwd())"
                )
                g = globals().copy()
                exec(patched, g)
                self.progress_signal.emit("🚀 Engine loaded — starting search...")
                self.progress_signal.emit("─" * 52)
                result = g["extract_local_leads"](
                    search_query      = self.search_query,
                    allowed_ratings   = self.allowed_ratings,
                    target_city       = self.target_city,
                    progress_callback = lambda m: self.progress_signal.emit(m)
                )
                self.finished_signal.emit(result)
            else:
                self.error_signal.emit(f"Server error: {response.status_code}")
        except Exception as e:
            self.error_signal.emit(f"Fetch failed: {str(e)}")


# ── Worker: Gmail auth check ──────────────────────────────────────
class GmailAuthWorker(QThread):
    success_signal = pyqtSignal(str)   # emits connected email address
    error_signal   = pyqtSignal(str)
    progress_signal = pyqtSignal(str)

    def __init__(self, base_dir, do_full_auth=False):
        super().__init__()
        self.base_dir      = base_dir
        self.do_full_auth  = do_full_auth   # True = trigger browser OAuth

    def run(self):
        try:
            URL = (f"https://raw.githubusercontent.com/skengeneral/TimeVehicle-basic"
                   f"/main/draft_engine.py?t={int(time.time())}")
            self.progress_signal.emit("🌐 Fetching draft engine...")
            resp = requests.get(URL, headers={'User-Agent':'Mozilla/5.0'}, timeout=15)
            if resp.status_code != 200:
                self.error_signal.emit(f"Could not fetch draft engine ({resp.status_code})")
                return
            g = globals().copy()
            exec(resp.text, g)

            if self.do_full_auth:
                self.progress_signal.emit("🌐 Opening Gmail sign-in in browser...")
                g["authenticate_gmail"](
                    base_dir          = self.base_dir,
                    progress_callback = lambda m: self.progress_signal.emit(m)
                )

            email = g["get_gmail_profile"](base_dir=self.base_dir)
            if email:
                self.success_signal.emit(email)
            else:
                self.error_signal.emit("Not connected — please click Connect Gmail")
        except Exception as e:
            self.error_signal.emit(str(e))


# ── Worker: bulk draft creation ───────────────────────────────────
class DraftWorker(QThread):
    finished_signal = pyqtSignal(dict)
    error_signal    = pyqtSignal(str)
    progress_signal = pyqtSignal(str)

    def __init__(self, base_dir):
        super().__init__()
        self.base_dir = base_dir

    def run(self):
        try:
            URL = (f"https://raw.githubusercontent.com/skengeneral/TimeVehicle-basic"
                   f"/main/draft_engine.py?t={int(time.time())}")
            self.progress_signal.emit("🌐 Fetching draft engine from cloud...")
            resp = requests.get(URL, headers={'User-Agent':'Mozilla/5.0'}, timeout=15)
            if resp.status_code != 200:
                self.error_signal.emit(f"Could not fetch draft engine ({resp.status_code})")
                return
            g = globals().copy()
            exec(resp.text, g)
            self.progress_signal.emit("✅ Draft engine loaded")
            self.progress_signal.emit("─" * 50)
            result = g["create_bulk_drafts"](
                base_dir          = self.base_dir,
                progress_callback = lambda m: self.progress_signal.emit(m)
            )
            self.finished_signal.emit(result)
        except Exception as e:
            self.error_signal.emit(str(e))


# ── Main UI ───────────────────────────────────────────────────────
class TimeVehicleUI(QWidget):
    def __init__(self):
        super().__init__()
        self.live_credits  = "Fetching..."
        self.scraper_worker = None
        self.draft_worker   = None
        self.gmail_worker   = None
        self.init_ui()
        QTimer.singleShot(100, self.lazy_load_credits)
        QTimer.singleShot(300, self.check_gmail_status_on_start)

    # ── Shared style constants ────────────────────────────────────
    BTN_PRIMARY = """
        QPushButton { background-color:#002D4A; color:white; padding:12px;
                      border:none; border-radius:4px; }
        QPushButton:hover    { background-color:#004473; }
        QPushButton:pressed  { background-color:#001524; }
        QPushButton:disabled { background-color:#64748B; color:#CBD5E1; }"""
    BTN_GREEN = """
        QPushButton { background-color:#10B981; color:white; border:none;
                      border-radius:4px; padding:6px 14px; }
        QPushButton:hover { background-color:#059669; }"""
    BTN_ORANGE = """
        QPushButton { background-color:#F59E0B; color:white; border:none;
                      border-radius:4px; padding:6px 14px; }
        QPushButton:hover { background-color:#D97706; }"""
    FRAME_BOX = ("background-color:#F8FAFC; border:1px solid #E2E8F0;"
                 "border-radius:6px;")
    LOG_STYLE  = ("QTextEdit { background-color:#0F1E2A; color:#90EE90;"
                  "border:none; border-radius:4px; padding:6px; }")

    def init_ui(self):
        self.setWindowTitle("Time Vehicle - 1.0")
        icon_path = os.path.join(_app_base(), "timevehicle.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self.setMinimumSize(480, 640)
        self.resize(500, 760)
        self.setStyleSheet("background-color:#FFFFFF;")

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # ── Shared header (above tabs) ────────────────────────────
        header = QWidget()
        header.setStyleSheet("background-color:#002D4A; border:none;")
        h_lay = QHBoxLayout(header)
        h_lay.setContentsMargins(25, 18, 25, 18)
        h_lay.setSpacing(14)
        h_lay.addWidget(LogoWidget())
        lbl = QLabel("TIME VEHICLE - 1.0")
        lbl.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        lbl.setStyleSheet("color:#FFFFFF;")
        h_lay.addWidget(lbl)
        h_lay.addStretch()
        outer.addWidget(header)

        # ── Tab widget ────────────────────────────────────────────
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane  { border:none; background:#FFFFFF; }
            QTabBar::tab      { background:#001F33; color:#90A4B4;
                                padding:11px 28px; font:bold 10px Arial;
                                border:none; }
            QTabBar::tab:selected { background:#002D4A; color:#FFFFFF;
                                    border-bottom:3px solid #10B981; }
            QTabBar::tab:hover    { background:#002D4A; color:#FFFFFF; }
        """)
        self.tabs.addTab(self._build_search_tab(), "🔍   SEARCH")
        self.tabs.addTab(self._build_drafts_tab(), "✉️   BULK DRAFTS")
        outer.addWidget(self.tabs)

    # ═════════════════════════════════════════════════════════════
    #  TAB 1 — SEARCH
    # ═════════════════════════════════════════════════════════════
    def _build_search_tab(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet("QScrollArea{border:none; background:#FFFFFF;}")

        content = QWidget()
        content.setStyleSheet("background:#FFFFFF;")
        lay = QVBoxLayout(content)
        lay.setSpacing(12)
        lay.setContentsMargins(25, 15, 25, 20)

        lf = QFont("Arial", 10)
        sf = QFont("Arial", 11, QFont.Weight.Bold)
        inp = ("padding:8px; border:1px solid #B0C4DE; border-radius:4px;"
               "background:#FFFFFF; color:#000000;")
        cb_style = ("QCheckBox{color:#333; padding:4px;}"
                    "QCheckBox::indicator{width:16px;height:16px;}")

        # Credits bar
        cf = QFrame(); cf.setStyleSheet(self.FRAME_BOX)
        cl = QHBoxLayout(cf); cl.setContentsMargins(12,6,12,6)
        self.lbl_credits = QLabel(f"💳 Searches Remaining: {self.live_credits}", font=lf)
        self.lbl_credits.setStyleSheet("color:#0F172A; font-weight:bold;")
        btn_rch = QPushButton("🔌 Recharge")
        btn_rch.setFont(QFont("Arial",9,QFont.Weight.Bold))
        btn_rch.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_rch.setStyleSheet(self.BTN_GREEN)
        btn_rch.clicked.connect(self.open_serpapi_billing)
        cl.addWidget(self.lbl_credits); cl.addStretch(); cl.addWidget(btn_rch)
        lay.addWidget(cf)

        # Profession
        lay.addWidget(self._lbl("Field of search / profession:", lf))
        self.input_profession = QLineEdit()
        self.input_profession.setPlaceholderText("e.g., Orthopedic doctor")
        self.input_profession.setStyleSheet(inp); lay.addWidget(self.input_profession)

        # Locality
        lay.addWidget(self._lbl("Locality / area name:", lf))
        self.input_locality = QLineEdit()
        self.input_locality.setPlaceholderText("e.g., Manhattan")
        self.input_locality.setStyleSheet(inp); lay.addWidget(self.input_locality)

        # City / State / Country
        geo = QHBoxLayout(); geo.setSpacing(15)
        for attr, lbl_txt, ph in [
            ("input_city",    "City:",    "e.g., New York"),
            ("input_state",   "State:",   "e.g., New York"),
            ("input_country", "Country:", "e.g., USA"),
        ]:
            vb = QVBoxLayout()
            vb.addWidget(self._lbl(lbl_txt, lf))
            field = QLineEdit(); field.setPlaceholderText(ph); field.setStyleSheet(inp)
            setattr(self, attr, field); vb.addWidget(field)
            geo.addLayout(vb, stretch=1)
        lay.addLayout(geo)

        self._sep(lay)

        # Ratings
        lay.addWidget(self._lbl("Google Rating Range", sf, "#002D4A"))
        rl = QHBoxLayout()
        self.rating_checkboxes = {}
        for rate in ["5","4","3","2","1","0","ALL"]:
            cb = QCheckBox(rate); cb.setFont(lf); cb.setStyleSheet(cb_style)
            if rate == "ALL":
                cb.setChecked(True)
                cb.toggled.connect(self.handle_all_ratings_toggle)
            else:
                cb.toggled.connect(self.handle_single_rating_toggle)
            rl.addWidget(cb); self.rating_checkboxes[rate] = cb
        lay.addLayout(rl)

        self._sep(lay)

        # Download format
        lay.addWidget(self._lbl("Download to System:", sf, "#002D4A"))
        fl = QHBoxLayout()
        self.chk_excel = QCheckBox("Excel"); self.chk_excel.setChecked(True)
        self.chk_excel.setFont(lf); self.chk_excel.setStyleSheet(cb_style)
        fl.addWidget(self.chk_excel); lay.addLayout(fl)

        self._sep(lay); lay.addSpacing(4)

        # Submit
        self.btn_submit = QPushButton("SUBMIT")
        self.btn_submit.setFont(QFont("Arial",12,QFont.Weight.Bold))
        self.btn_submit.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_submit.setStyleSheet(self.BTN_PRIMARY)
        self.btn_submit.clicked.connect(self.handle_submit_action)
        lay.addWidget(self.btn_submit)

        # Progress log
        self.progress_frame = QFrame()
        self.progress_frame.setStyleSheet(
            "QFrame{background:#F0F4F8;border:1px solid #B0C4DE;border-radius:6px;}")
        pf_lay = QVBoxLayout(self.progress_frame)
        pf_lay.setContentsMargins(10,8,10,8); pf_lay.setSpacing(5)
        ph = QHBoxLayout()
        pt = QLabel("📊  Live Progress Log")
        pt.setFont(QFont("Arial",10,QFont.Weight.Bold))
        pt.setStyleSheet("color:#002D4A;border:none;background:transparent;")
        self.lbl_lead_count = QLabel("Leads collected: 0")
        self.lbl_lead_count.setFont(QFont("Arial",10,QFont.Weight.Bold))
        self.lbl_lead_count.setStyleSheet("color:#059669;border:none;background:transparent;")
        ph.addWidget(pt); ph.addStretch(); ph.addWidget(self.lbl_lead_count)
        pf_lay.addLayout(ph)
        self.progress_log = QTextEdit()
        self.progress_log.setReadOnly(True)
        self.progress_log.setFont(QFont("Consolas",9))
        self.progress_log.setFixedHeight(190)
        self.progress_log.setStyleSheet(self.LOG_STYLE)
        pf_lay.addWidget(self.progress_log)
        self.progress_frame.setVisible(False)
        lay.addWidget(self.progress_frame)

        lay.addSpacing(15)

        # Support footer
        sf_frame = QFrame()
        sf_frame.setStyleSheet(self.FRAME_BOX)
        sf_lay = QHBoxLayout(sf_frame)
        sup = QLabel("💬 Need help or customised solutions? WhatsApp: +91 77803 79259\n"
                     "📧 support@timevehicle.com")
        sup.setFont(QFont("Arial",10,QFont.Weight.Bold))
        sup.setStyleSheet("color:#475569; line-height:1.4;")
        sup.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sf_lay.addWidget(sup); lay.addWidget(sf_frame)

        scroll.setWidget(content)
        return scroll

    # ═════════════════════════════════════════════════════════════
    #  TAB 2 — BULK DRAFTS
    # ═════════════════════════════════════════════════════════════
    def _build_drafts_tab(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet("QScrollArea{border:none; background:#FFFFFF;}")

        content = QWidget(); content.setStyleSheet("background:#FFFFFF;")
        lay = QVBoxLayout(content)
        lay.setSpacing(14); lay.setContentsMargins(25, 18, 25, 20)

        lf = QFont("Arial", 10)
        bf = QFont("Arial", 10, QFont.Weight.Bold)
        sf = QFont("Arial", 11, QFont.Weight.Bold)

        # ── 1. Leads file status ──────────────────────────────────
        leads_frame = QFrame(); leads_frame.setStyleSheet(self.FRAME_BOX)
        ll = QVBoxLayout(leads_frame); ll.setContentsMargins(14,10,14,10); ll.setSpacing(4)
        lbl_leads_title = QLabel("📄  Leads File Status")
        lbl_leads_title.setFont(sf); lbl_leads_title.setStyleSheet("color:#002D4A;")
        self.lbl_leads_file = QLabel("No search completed yet — run a search in Tab 1 first")
        self.lbl_leads_file.setFont(lf); self.lbl_leads_file.setStyleSheet("color:#64748B;")
        self.lbl_selected_count = QLabel("")
        self.lbl_selected_count.setFont(bf)
        self.lbl_selected_count.setStyleSheet("color:#059669;")
        ll.addWidget(lbl_leads_title)
        ll.addWidget(self.lbl_leads_file)
        ll.addWidget(self.lbl_selected_count)
        lay.addWidget(leads_frame)

        self._sep(lay)

        # ── 2. Anthropic API key ──────────────────────────────────
        api_frame = QFrame(); api_frame.setStyleSheet(self.FRAME_BOX)
        al = QVBoxLayout(api_frame); al.setContentsMargins(14,10,14,10); al.setSpacing(6)
        lbl_api_title = QLabel("🔑  Anthropic API Key  (for AI email rewriting)")
        lbl_api_title.setFont(sf); lbl_api_title.setStyleSheet("color:#002D4A;")
        al.addWidget(lbl_api_title)
        api_row = QHBoxLayout()
        self.lbl_api_status = QLabel("Checking...")
        self.lbl_api_status.setFont(bf)
        self.lbl_api_status.setStyleSheet("color:#64748B;")
        btn_get_key = QPushButton("Get API Key →")
        btn_get_key.setFont(QFont("Arial",9,QFont.Weight.Bold))
        btn_get_key.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_get_key.setStyleSheet(self.BTN_ORANGE)
        btn_get_key.clicked.connect(lambda: webbrowser.open("https://console.anthropic.com/"))
        api_row.addWidget(self.lbl_api_status)
        api_row.addStretch()
        api_row.addWidget(btn_get_key)
        al.addLayout(api_row)
        lbl_api_hint = QLabel("Add your key to  anthropic_api.txt  in the same folder as this app")
        lbl_api_hint.setFont(QFont("Arial",9)); lbl_api_hint.setStyleSheet("color:#94A3B8;")
        al.addWidget(lbl_api_hint)
        lay.addWidget(api_frame)

        self._sep(lay)

        # ── 3. Gmail account ──────────────────────────────────────
        gmail_frame = QFrame(); gmail_frame.setStyleSheet(self.FRAME_BOX)
        gl = QVBoxLayout(gmail_frame); gl.setContentsMargins(14,10,14,10); gl.setSpacing(6)
        lbl_gmail_title = QLabel("📧  Gmail Account")
        lbl_gmail_title.setFont(sf); lbl_gmail_title.setStyleSheet("color:#002D4A;")
        gl.addWidget(lbl_gmail_title)
        gmail_row = QHBoxLayout()
        self.lbl_gmail_status = QLabel("Not connected")
        self.lbl_gmail_status.setFont(bf)
        self.lbl_gmail_status.setStyleSheet("color:#64748B;")
        self.btn_connect_gmail = QPushButton("Connect Gmail")
        self.btn_connect_gmail.setFont(QFont("Arial",9,QFont.Weight.Bold))
        self.btn_connect_gmail.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_connect_gmail.setStyleSheet(self.BTN_GREEN)
        self.btn_connect_gmail.clicked.connect(self.on_connect_gmail)
        gmail_row.addWidget(self.lbl_gmail_status)
        gmail_row.addStretch()
        gmail_row.addWidget(self.btn_connect_gmail)
        gl.addLayout(gmail_row)
        lbl_gmail_hint = QLabel("First time only — a browser popup will ask you to sign in")
        lbl_gmail_hint.setFont(QFont("Arial",9)); lbl_gmail_hint.setStyleSheet("color:#94A3B8;")
        gl.addWidget(lbl_gmail_hint)
        lay.addWidget(gmail_frame)

        self._sep(lay)

        # ── 4. Create Bulk Drafts button ──────────────────────────
        self.btn_create_drafts = QPushButton("✉️   CREATE BULK DRAFTS")
        self.btn_create_drafts.setFont(QFont("Arial",12,QFont.Weight.Bold))
        self.btn_create_drafts.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_create_drafts.setStyleSheet(self.BTN_PRIMARY)
        self.btn_create_drafts.clicked.connect(self.on_create_drafts)
        lay.addWidget(self.btn_create_drafts)

        # ── 5. Draft progress log ─────────────────────────────────
        self.draft_progress_frame = QFrame()
        self.draft_progress_frame.setStyleSheet(
            "QFrame{background:#F0F4F8;border:1px solid #B0C4DE;border-radius:6px;}")
        dpf = QVBoxLayout(self.draft_progress_frame)
        dpf.setContentsMargins(10,8,10,8); dpf.setSpacing(5)
        dp_head = QHBoxLayout()
        dp_title = QLabel("📊  Live Progress Log")
        dp_title.setFont(QFont("Arial",10,QFont.Weight.Bold))
        dp_title.setStyleSheet("color:#002D4A;border:none;background:transparent;")
        self.lbl_draft_count = QLabel("Drafts created: 0")
        self.lbl_draft_count.setFont(QFont("Arial",10,QFont.Weight.Bold))
        self.lbl_draft_count.setStyleSheet("color:#059669;border:none;background:transparent;")
        dp_head.addWidget(dp_title); dp_head.addStretch(); dp_head.addWidget(self.lbl_draft_count)
        dpf.addLayout(dp_head)
        self.draft_log = QTextEdit()
        self.draft_log.setReadOnly(True)
        self.draft_log.setFont(QFont("Consolas",9))
        self.draft_log.setFixedHeight(200)
        self.draft_log.setStyleSheet(self.LOG_STYLE)
        dpf.addWidget(self.draft_log)
        self.draft_progress_frame.setVisible(False)
        lay.addWidget(self.draft_progress_frame)

        lay.addStretch()
        scroll.setWidget(content)
        return scroll

    # ═════════════════════════════════════════════════════════════
    #  SHARED HELPERS
    # ═════════════════════════════════════════════════════════════
    def _lbl(self, text, font, color="#333333"):
        l = QLabel(text, font=font); l.setStyleSheet(f"color:{color};"); return l

    def _sep(self, layout):
        line = QFrame(); line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet("background-color:#E2E8F0; height:1px; border:none;"
                           "margin-top:4px; margin-bottom:4px;")
        layout.addWidget(line)

    def add_separator(self, layout):
        self._sep(layout)

    # ═════════════════════════════════════════════════════════════
    #  TAB 1 LOGIC
    # ═════════════════════════════════════════════════════════════
    def handle_all_ratings_toggle(self, checked):
        if checked:
            for rate, cb in self.rating_checkboxes.items():
                if rate != "ALL":
                    cb.blockSignals(True); cb.setChecked(False); cb.blockSignals(False)

    def handle_single_rating_toggle(self, checked):
        if checked:
            self.rating_checkboxes["ALL"].blockSignals(True)
            self.rating_checkboxes["ALL"].setChecked(False)
            self.rating_checkboxes["ALL"].blockSignals(False)

    def open_serpapi_billing(self, _=None):
        webbrowser.open("https://serpapi.com/plan")
        self.refresh_credit_display()

    def lazy_load_credits(self):
        self.live_credits = fetch_live_serp_credits()
        self.lbl_credits.setText(f"💳 Searches Remaining: {self.live_credits}")

    def refresh_credit_display(self):
        self.lbl_credits.setText("💳 Searches Remaining: Fetching...")
        self.live_credits = fetch_live_serp_credits()
        self.lbl_credits.setText(f"💳 Searches Remaining: {self.live_credits}")

    def verify_security_passkey(self):
        saved = _load_saved_passkey()
        if saved:
            result = _validate_passkey_cloud(saved)
            if result is True:  return True
            if result is None:  return True
            _clear_saved_passkey()
            QMessageBox.warning(self, "Access Key Expired",
                "Your Time Vehicle access key has expired or been revoked.\n"
                "Please enter a new passkey to continue.")

        passkey, ok = QInputDialog.getText(
            self, "Security Verification",
            "Please enter your Time Vehicle Activation Passkey:",
            QLineEdit.EchoMode.Normal)
        if not ok: return False
        pk = passkey.strip()
        if not pk:
            QMessageBox.warning(self, "Empty", "Passkey cannot be empty."); return False

        result = _validate_passkey_cloud(pk)
        if result is True:
            _save_passkey(pk); return True
        if result is None:
            QMessageBox.critical(self, "Network Failure",
                "Could not reach validation server.\nCheck your internet and try again.")
            return False
        QMessageBox.critical(self, "Verification Failure",
            "Passkey not recognised or has expired.")
        return False

    def handle_submit_action(self):
        profession = self.input_profession.text().strip()
        city       = self.input_city.text().strip()
        locality   = self.input_locality.text().strip()
        state      = self.input_state.text().strip()
        country    = self.input_country.text().strip()

        if not profession or not city:
            QMessageBox.warning(self, "Missing Info",
                "Profession and City are required fields."); return

        selected_ratings = (["ALL"] if self.rating_checkboxes["ALL"].isChecked()
                            else [r for r, cb in self.rating_checkboxes.items()
                                  if r != "ALL" and cb.isChecked()])
        if not selected_ratings:
            QMessageBox.warning(self, "Missing Rating",
                "Please select at least one rating or ALL."); return

        if not self.verify_security_passkey(): return

        # Browser install
        self.btn_submit.setText("⚙️ PREPARING BROWSER...")
        self.btn_submit.setEnabled(False); self.btn_submit.repaint()
        try:
            if getattr(sys, 'frozen', False):
                from playwright._impl._driver import compute_driver_executable
                subprocess.check_call([str(compute_driver_executable()), "install", "chromium"])
            else:
                subprocess.check_call([sys.executable, "-m", "playwright", "install", "chromium"])
        except Exception as e:
            QMessageBox.warning(self, "Setup Failed",
                f"Browser setup failed. Check internet.\n{e}")
            self.btn_submit.setText("SUBMIT"); self.btn_submit.setEnabled(True); return

        query  = " ".join(filter(None, [profession, locality]))
        target = ", ".join(filter(None, [city, state, country]))

        self.progress_log.clear()
        self.lbl_lead_count.setText("Leads collected: 0")
        self.progress_frame.setVisible(True)
        self.btn_submit.setText("⏳ DOWNLOADING DATA...")

        self.scraper_worker = ScraperWorker(query, selected_ratings, target)
        self.scraper_worker.finished_signal.connect(self.on_scraping_finished)
        self.scraper_worker.error_signal.connect(self.on_scraping_error)
        self.scraper_worker.progress_signal.connect(self.on_progress_update)
        self.scraper_worker.start()

    def on_progress_update(self, message):
        self.progress_log.append(message)
        sb = self.progress_log.verticalScrollBar(); sb.setValue(sb.maximum())
        if message.startswith("🏢 ["):
            try:
                count = int(message.split("[")[1].split("]")[0])
                self.lbl_lead_count.setText(f"Leads collected: {count}")
            except: pass
        if "DONE —" in message:
            try:
                total = int(message.split("DONE —")[1].split("qualified")[0].strip())
                self.lbl_lead_count.setText(f"✅ Total leads: {total}")
            except: pass

    def on_scraping_finished(self, packet):
        try:
            data = packet.get("data", [])
            if not data:
                QMessageBox.warning(self, "No Results",
                    "No businesses found. Try broadening your search.")
                return
            if self.chk_excel.isChecked():
                URL = ("https://raw.githubusercontent.com/skengeneral/"
                       "TimeVehicle-basic/main/export_engine.py")
                resp = requests.get(URL, headers={'User-Agent':'Mozilla/5.0'}, timeout=10)
                if resp.status_code == 200:
                    code = (resp.text
                            .replace("from openpyxl import Workbook","")
                            .replace("from openpyxl.styles import Font, PatternFill, Alignment, Border, Side","")
                            .replace("from openpyxl.worksheet.datavalidation import DataValidation","")
                            .replace("os.path.abspath(__file__)","os.getcwd()")
                            .replace("__file__","os.path.abspath(os.getcwd())"))
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.worksheet.datavalidation import DataValidation
                    scope = {
                        "__builtins__": __builtins__,
                        "Workbook": Workbook, "Font": Font,
                        "PatternFill": PatternFill, "Alignment": Alignment,
                        "Border": Border, "Side": Side,
                        "DataValidation": DataValidation,
                        "os": os, "sys": sys,
                    }
                    exec(code, scope)
                    xl_path = scope["save_to_excel"](data, packet.get("columns_layout"))
                    # Update Tab 2 leads status
                    self.refresh_leads_status()
                    self.refresh_credit_display()
                    QMessageBox.information(self, "✅ Search Complete",
                        f"Total leads: {len(data)}\n\nSaved to:\n{xl_path}\n\n"
                        f"Open the file, select emails, fill Subject & Body,\n"
                        f"then switch to the ✉️ BULK DRAFTS tab to send.")
                else:
                    raise Exception(f"Export engine returned {resp.status_code}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))
        finally:
            self.btn_submit.setEnabled(True); self.btn_submit.setText("SUBMIT")

    def on_scraping_error(self, msg):
        QMessageBox.critical(self, "Search Failed", msg)
        self.btn_submit.setEnabled(True); self.btn_submit.setText("SUBMIT")

    # ═════════════════════════════════════════════════════════════
    #  TAB 2 LOGIC
    # ═════════════════════════════════════════════════════════════
    def check_gmail_status_on_start(self):
        """Silent check on app launch — update status if already authorised."""
        self._refresh_api_key_status()
        worker = GmailAuthWorker(_app_base(), do_full_auth=False)
        worker.success_signal.connect(self._on_gmail_connected)
        worker.error_signal.connect(lambda _: None)   # silent on start
        worker.progress_signal.connect(lambda _: None)
        worker.start()

    def refresh_leads_status(self):
        """Called after a search completes to update Tab 2 leads panel."""
        leads_path = os.path.join(_app_base(), "Time_Vehicle_Leads.xlsx")
        if not os.path.exists(leads_path):
            self.lbl_leads_file.setText("No leads file found — run a search first")
            self.lbl_leads_file.setStyleSheet("color:#64748B;")
            self.lbl_selected_count.setText("")
            return
        try:
            from openpyxl import load_workbook
            wb  = load_workbook(leads_path, read_only=True)
            ws  = wb.active
            rows = list(ws.iter_rows(values_only=True))
            total    = sum(1 for r in rows[1:] if r and r[0] not in
                           (None,"","MAIL SUBJECT","MAIL BODY\nTEMPLATE","MAIL BODY TEMPLATE")
                           and "DRAFT EMAIL" not in str(r[0]))
            selected = sum(1 for r in rows[1:] if r and str(r[0]).strip().upper() == "YES")
            self.lbl_leads_file.setText(f"✅  Time_Vehicle_Leads.xlsx  —  {total} total leads")
            self.lbl_leads_file.setStyleSheet("color:#059669;")
            self.lbl_selected_count.setText(
                f"{'✉️  ' + str(selected) + ' emails selected for drafting' if selected else '⚠️  No emails selected yet — open the file and set SELECT = Yes'}"
            )
            self.lbl_selected_count.setStyleSheet(
                "color:#059669;" if selected else "color:#F59E0B;")
        except Exception as e:
            self.lbl_leads_file.setText(f"Could not read leads file: {e}")

    def _refresh_api_key_status(self):
        found, masked = get_anthropic_key_status()
        if found:
            self.lbl_api_status.setText(f"✅  Key found:  {masked}")
            self.lbl_api_status.setStyleSheet("color:#059669; font-weight:bold;")
        else:
            self.lbl_api_status.setText("❌  Key not found — add to anthropic_api.txt")
            self.lbl_api_status.setStyleSheet("color:#EF4444; font-weight:bold;")

    def on_connect_gmail(self):
        self.btn_connect_gmail.setEnabled(False)
        self.btn_connect_gmail.setText("Connecting...")
        self.lbl_gmail_status.setText("Opening browser for sign-in...")
        self.lbl_gmail_status.setStyleSheet("color:#F59E0B; font-weight:bold;")

        self.gmail_worker = GmailAuthWorker(_app_base(), do_full_auth=True)
        self.gmail_worker.success_signal.connect(self._on_gmail_connected)
        self.gmail_worker.error_signal.connect(self._on_gmail_error)
        self.gmail_worker.progress_signal.connect(
            lambda m: self.lbl_gmail_status.setText(m))
        self.gmail_worker.start()

    def _on_gmail_connected(self, email):
        self.lbl_gmail_status.setText(f"✅  Connected:  {email}")
        self.lbl_gmail_status.setStyleSheet("color:#059669; font-weight:bold;")
        self.btn_connect_gmail.setText("Reconnect")
        self.btn_connect_gmail.setEnabled(True)

    def _on_gmail_error(self, err):
        self.lbl_gmail_status.setText(f"❌  {err[:80]}")
        self.lbl_gmail_status.setStyleSheet("color:#EF4444; font-weight:bold;")
        self.btn_connect_gmail.setText("Connect Gmail")
        self.btn_connect_gmail.setEnabled(True)

    def on_create_drafts(self):
        # Refresh status first
        self._refresh_api_key_status()
        self.refresh_leads_status()

        found_key, _ = get_anthropic_key_status()
        if not found_key:
            QMessageBox.warning(self, "API Key Missing",
                "Please add your Anthropic API key to anthropic_api.txt first.")
            return

        leads_path = os.path.join(_app_base(), "Time_Vehicle_Leads.xlsx")
        if not os.path.exists(leads_path):
            QMessageBox.warning(self, "No Leads File",
                "Please run a search first to generate Time_Vehicle_Leads.xlsx.")
            return

        self.draft_log.clear()
        self.lbl_draft_count.setText("Drafts created: 0")
        self.draft_progress_frame.setVisible(True)
        self.btn_create_drafts.setEnabled(False)
        self.btn_create_drafts.setText("⏳  CREATING DRAFTS...")

        self.draft_worker = DraftWorker(_app_base())
        self.draft_worker.finished_signal.connect(self.on_drafts_finished)
        self.draft_worker.error_signal.connect(self.on_drafts_error)
        self.draft_worker.progress_signal.connect(self.on_draft_progress)
        self.draft_worker.start()

    def on_draft_progress(self, msg):
        self.draft_log.append(msg)
        sb = self.draft_log.verticalScrollBar(); sb.setValue(sb.maximum())
        if "Draft created" in msg or "✅ Draft" in msg:
            try:
                current = int(self.lbl_draft_count.text().split(":")[1].strip().split("/")[0])
                self.lbl_draft_count.setText(f"✉️  Drafts created: {current + 1}")
            except: pass
        if "DONE —" in msg:
            try:
                n = int(msg.split("DONE —")[1].split("draft")[0].strip())
                self.lbl_draft_count.setText(f"✅  {n} drafts created")
            except: pass

    def on_drafts_finished(self, result):
        created = result.get("created", 0)
        failed  = result.get("failed",  0)
        total   = result.get("total",   0)
        self.lbl_draft_count.setText(f"✅  {created} / {total} drafts created")
        self.btn_create_drafts.setEnabled(True)
        self.btn_create_drafts.setText("✉️   CREATE BULK DRAFTS")
        msg = (f"{created} draft{'s' if created!=1 else ''} created in your Gmail.\n\n"
               f"Open Gmail → Drafts to review and send them.")
        if failed:
            msg += f"\n\n⚠️ {failed} failed — check the progress log for details."
        QMessageBox.information(self, "✅ Drafts Created", msg)

    def on_drafts_error(self, err):
        QMessageBox.critical(self, "Draft Creation Failed", err)
        self.btn_create_drafts.setEnabled(True)
        self.btn_create_drafts.setText("✉️   CREATE BULK DRAFTS")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    if sys.platform == "win32":
        app.setStyle("Fusion")
    win = TimeVehicleUI()
    win.show()
    sys.exit(app.exec())
