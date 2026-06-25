import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def save_to_excel(data, columns_layout=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Local Business Leads"

    ws.views.sheetView[0].showGridLines = True

    # ── Colour palette ────────────────────────────────────────────
    header_fill  = PatternFill(start_color="002D4A", end_color="002D4A", fill_type="solid")
    select_fill  = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    select_cell_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    section_fill = PatternFill(start_color="001F33", end_color="001F33", fill_type="solid")
    input_fill   = PatternFill(start_color="FDFEFE", end_color="FDFEFE", fill_type="solid")

    header_font  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font    = Font(name="Arial", size=10, color="000000")
    section_font = Font(name="Arial", size=11, bold=True, color="00E5CC")
    input_font   = Font(name="Arial", size=10, color="000000")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    top_left     = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    thin         = Side(border_style="thin", color="E2E8F0")
    thick        = Side(border_style="medium", color="002D4A")
    cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    # ── Headers — SELECT is always column A ───────────────────────
    data_headers = [
        "Business Name", "Google Rating", "Complete Address",
        "Operating Hours Matrix", "Website Link", "Email ID",
        "Phone Number", "Facebook Handle", "Instagram Handle",
        "LinkedIn Handle", "Twitter/X Handle"
    ]
    all_headers = ["SELECT"] + data_headers
    total_cols  = len(all_headers)

    # Write header row
    ws.append(all_headers)
    for col_num, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill   = select_fill if header == "SELECT" else header_fill
        cell.font   = header_font
        cell.alignment = center_align
        cell.border = cell_border
    ws.row_dimensions[1].height = 28

    # ── Yes / No dropdown validation for SELECT column ────────────
    dv = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=False,
        showDropDown=False,       # False = show the ▼ arrow
        showErrorMessage=True,
        errorTitle="Invalid",
        error="Please choose Yes or No from the dropdown."
    )
    ws.add_data_validation(dv)

    # ── Data rows ─────────────────────────────────────────────────
    for row_idx, item in enumerate(data, 2):
        row_values = ["No"] + [item.get(h, "Not Provided") for h in data_headers]
        ws.append(row_values)
        ws.row_dimensions[row_idx].height = 22

        for col_idx, header in enumerate(all_headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font   = data_font
            cell.border = cell_border
            if header == "SELECT":
                cell.fill      = select_cell_fill
                cell.alignment = center_align
                dv.add(cell)                      # attach dropdown
            elif header == "Google Rating":
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # ── Column widths ─────────────────────────────────────────────
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len    = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
    ws.column_dimensions["A"].width = 10   # SELECT column fixed narrow

    # ── Draft email settings section ──────────────────────────────
    last_data_row = len(data) + 1
    section_row   = last_data_row + 2      # one blank gap row
    subject_row   = section_row + 1
    body_row      = subject_row + 1

    # Section banner
    banner = ws.cell(row=section_row, column=1,
                     value="✉️   DRAFT EMAIL SETTINGS  —  Fill in Subject & Body below, then open draft_engine.exe")
    banner.fill      = section_fill
    banner.font      = section_font
    banner.alignment = left_align
    banner.border    = section_border
    ws.merge_cells(start_row=section_row, start_column=1,
                   end_row=section_row,   end_column=total_cols)
    ws.row_dimensions[section_row].height = 26

    # MAIL SUBJECT label + input
    subj_lbl = ws.cell(row=subject_row, column=1, value="MAIL SUBJECT")
    subj_lbl.fill      = header_fill
    subj_lbl.font      = header_font
    subj_lbl.alignment = center_align
    subj_lbl.border    = cell_border

    subj_inp = ws.cell(row=subject_row, column=2, value="")
    subj_inp.fill      = input_fill
    subj_inp.font      = input_font
    subj_inp.alignment = left_align
    subj_inp.border    = cell_border
    ws.merge_cells(start_row=subject_row, start_column=2,
                   end_row=subject_row,   end_column=total_cols)
    ws.row_dimensions[subject_row].height = 26

    # MAIL BODY TEMPLATE label + input
    body_lbl = ws.cell(row=body_row, column=1, value="MAIL BODY\nTEMPLATE")
    body_lbl.fill      = header_fill
    body_lbl.font      = header_font
    body_lbl.alignment = center_align
    body_lbl.border    = cell_border

    body_inp = ws.cell(row=body_row, column=2, value="")
    body_inp.fill      = input_fill
    body_inp.font      = input_font
    body_inp.alignment = top_left          # top-align for multi-line
    body_inp.border    = cell_border
    ws.merge_cells(start_row=body_row, start_column=2,
                   end_row=body_row,   end_column=total_cols)
    ws.row_dimensions[body_row].height = 150   # tall for multi-line body

    # ── Save ──────────────────────────────────────────────────────
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))

    filename = os.path.join(base_path, "Time_Vehicle_Leads.xlsx")
    wb.save(filename)
    return filename
