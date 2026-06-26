import os
import sys
import base64
import time
import requests
from pathlib import Path
from email.mime.text import MIMEText
from openpyxl import load_workbook

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.compose']

# ── Path helpers ──────────────────────────────────────────────────
def _base(base_dir=None):
    if base_dir:
        return Path(base_dir)
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(os.getcwd())

def get_anthropic_key(base_dir=None):
    f = _base(base_dir) / "anthropic_api.txt"
    if f.exists():
        return f.read_text(encoding="utf-8").strip()
    return os.environ.get("ANTHROPIC_API_KEY", "")

def get_credentials_path(base_dir=None):
    return _base(base_dir) / "credentials.json"

def get_token_path(base_dir=None):
    return _base(base_dir) / "gmail_token.json"

# ── Gmail auth ────────────────────────────────────────────────────
def authenticate_gmail(base_dir=None, progress_callback=None):
    def log(m):
        if progress_callback: progress_callback(m)

    creds_path = get_credentials_path(base_dir)
    token_path = get_token_path(base_dir)

    if not creds_path.exists():
        raise FileNotFoundError(
            "credentials.json not found.\n"
            "Please contact Time Vehicle support to obtain this file."
        )

    creds = None
    if token_path.exists():
        try:
            creds = Credentials.from_authorized_user_file(str(token_path), GMAIL_SCOPES)
        except Exception:
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            log("🔄 Refreshing Gmail session...")
            creds.refresh(Request())
        else:
            log("🌐 Opening Gmail sign-in in your browser...")
            flow = InstalledAppFlow.from_client_secrets_file(
                str(creds_path), GMAIL_SCOPES
            )
            creds = flow.run_local_server(port=0, open_browser=True)
            log("✅ Gmail authorization granted")

        token_path.write_text(creds.to_json())

    return build("gmail", "v1", credentials=creds)

def get_gmail_profile(base_dir=None):
    token_path = get_token_path(base_dir)
    creds_path = get_credentials_path(base_dir)
    if not token_path.exists() or not creds_path.exists():
        return None
    try:
        creds = Credentials.from_authorized_user_file(str(token_path), GMAIL_SCOPES)
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        if creds and creds.valid:
            svc     = build("gmail", "v1", credentials=creds)
            profile = svc.users().getProfile(userId="me").execute()
            return profile.get("emailAddress")
    except Exception:
        pass
    return None

# ── Manual mode placeholder replacer ─────────────────────────────
def apply_manual_template(body_template, lead):
    """
    Replaces (Placeholder) tags in the body template with actual
    lead data from Excel. No Claude API call — zero cost.

    Supported placeholders:
        (Business Name)     → lead name
        (Complete Address)  → full address
        (Google Rating)     → star rating
        (Phone Number)      → phone
        (Website Link)      → website URL
        (Email ID)          → email address
    """
    mapping = {
        "(Business Name)":    lead.get("name",    ""),
        "(Complete Address)": lead.get("address", ""),
        "(Google Rating)":    lead.get("rating",  ""),
        "(Phone Number)":     lead.get("phone",   ""),
        "(Website Link)":     lead.get("website", ""),
        "(Email ID)":         lead.get("email",   ""),
    }
    result = body_template
    for placeholder, value in mapping.items():
        if value and value not in ("Not Provided", "No Website", ""):
            result = result.replace(placeholder, value)
    return result


# ── Claude body rewriter ──────────────────────────────────────────
def rewrite_body(api_key, body_template, lead):
    """
    Rewrites the body template uniquely for each business.
    Uses ONLY the data from the Excel file — never outside knowledge.
    """
    name    = lead.get("name",    "")
    rating  = lead.get("rating",  "")
    address = lead.get("address", "")
    phone   = lead.get("phone",   "")
    website = lead.get("website", "")

    # Build recipient context from available data
    context_lines = [f"Business name: {name}"]
    if rating and rating not in ("", "Not Provided", "0"):
        context_lines.append(f"Google Rating: {rating} ★")
    if address and address not in ("", "Not Provided"):
        context_lines.append(f"Address: {address}")
    if phone and phone not in ("", "Not Provided"):
        context_lines.append(f"Phone: {phone}")
    if website and website not in ("", "Not Provided", "No Website"):
        context_lines.append(f"Website: {website}")
    context = "\n".join(context_lines)

    prompt = (
        f"You are helping a professional send personalised outreach emails.\n\n"
        f"Recipient business details (from verified data — use ONLY these details):\n"
        f"{context}\n\n"
        f"Original email body template:\n---\n{body_template}\n---\n\n"
        f"STRICT RULES — you must follow all of these without exception:\n"
        f"1. Use ONLY the business details provided above — never use your own knowledge about this business\n"
        f"2. If you reference the business location, use ONLY the exact address provided: '{address}'\n"
        f"3. Do NOT invent, assume, or add any details not explicitly given above\n"
        f"4. If a detail is not provided, simply do not mention it\n"
        f"5. When referencing location, always use the format 'your [suburb/city] location' "
        f"(e.g. 'your Hatfield location') — never infer or describe the business type from the address\n"
        f"6. Naturally mention '{name}' once or twice where it fits organically\n"
        f"7. Vary sentence structure, vocabulary, and phrasing from the original template\n"
        f"8. Keep the same core message, intent, tone, and approximate length\n"
        f"9. Sound professional and human — not AI-generated\n"
        f"10. Body paragraphs only — no greeting, no sign-off, no subject line\n"
        f"11. Return ONLY the rewritten body text, nothing else"
    )

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "Content-Type": "application/json",
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01"
        },
        json={
            "model": "claude-sonnet-4-6",
            "max_tokens": 1000,
            "messages": [{"role": "user", "content": prompt}]
        },
        timeout=30
    )
    if resp.status_code == 200:
        return resp.json()["content"][0]["text"].strip()
    raise Exception(f"Claude API error {resp.status_code}: {resp.text[:200]}")

# ── Gmail draft creator ───────────────────────────────────────────
def create_draft(service, to_email, subject, body):
    msg            = MIMEText(body, "plain", "utf-8")
    msg["to"]      = to_email
    msg["subject"] = subject
    raw            = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    draft          = service.users().drafts().create(
        userId="me", body={"message": {"raw": raw}}
    ).execute()
    return draft.get("id")

# ── Leads reader ──────────────────────────────────────────────────
def read_leads(base_dir=None, email_selection="ALL", progress_callback=None):
    """
    Reads Time_Vehicle_Leads.xlsx and returns selected leads with
    full business details for rich Claude personalisation.

    email_selection:
        "ALL"           → every row with a valid email
        list of strings → only rows whose Email ID is in the list
    """
    def log(m):
        if progress_callback: progress_callback(m)

    leads_file = _base(base_dir) / "Time_Vehicle_Leads.xlsx"
    if not leads_file.exists():
        raise FileNotFoundError(
            "Time_Vehicle_Leads.xlsx not found.\n"
            "Please run a search first."
        )

    wb   = load_workbook(str(leads_file))
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise ValueError("Leads file appears empty. Please run a new search.")

    # Pull subject and body from the bottom section
    subject       = ""
    body_template = ""
    for row in rows:
        if not row or not row[0]: continue
        label = str(row[0]).strip().upper()
        if "MAIL SUBJECT" in label and len(row) > 1:
            subject = str(row[1] or "").strip()
        elif "MAIL BODY" in label and len(row) > 1:
            body_template = str(row[1] or "").strip()

    if not subject:
        raise ValueError(
            "MAIL SUBJECT is empty.\n"
            "Please fill it in Time_Vehicle_Leads.xlsx and save."
        )
    if not body_template:
        raise ValueError(
            "MAIL BODY TEMPLATE is empty.\n"
            "Please fill it in Time_Vehicle_Leads.xlsx and save."
        )

    # Build filter set if specific emails selected
    filter_emails = None
    if email_selection != "ALL":
        filter_emails = set(e.strip().lower() for e in email_selection if e.strip())

    # Extract lead rows with full business data
    header_row = rows[0]
    leads      = []

    for row in rows[1:]:
        if not row or not row[0]: continue
        label = str(row[0]).strip().upper()
        # Skip section header rows
        if any(x in label for x in ["MAIL SUBJECT", "MAIL BODY", "DRAFT EMAIL",
                                     "✉️", "TEMPLATE"]):
            continue

        lead    = {str(h): str(v or "").strip() for h, v in zip(header_row, row) if h}
        email   = lead.get("Email ID", "").strip()
        name    = lead.get("Business Name", "").strip()

        if not email or email.lower() == "not provided" or "@" not in email:
            continue

        # Apply email filter if specific selection
        if filter_emails is not None and email.lower() not in filter_emails:
            continue

        leads.append({
            "name":    name,
            "email":   email,
            "rating":  lead.get("Google Rating",       ""),
            "address": lead.get("Complete Address",     ""),
            "phone":   lead.get("Phone Number",         ""),
            "website": lead.get("Website Link",         ""),
            "hours":   lead.get("Operating Hours Matrix",""),
        })

    log(f"📋 {len(leads)} emails queued for drafting")
    log(f"📝 Subject: {subject}")
    return leads, subject, body_template

# ── Main entry point ──────────────────────────────────────────────
def create_bulk_drafts(base_dir=None, email_selection="ALL", draft_mode="AI", progress_callback=None):
    """
    Called from Tab 2 DraftWorker.
    draft_mode: "AI"     → Claude rewrites body uniquely per email
                "MANUAL" → template used as-is with (Placeholder) replacements
    """
    def log(m):
        if progress_callback: progress_callback(m)

    # 1. API key — only needed for AI mode
    api_key = None
    if draft_mode == "AI":
        api_key = get_anthropic_key(base_dir)
        if not api_key:
            raise ValueError(
                "Anthropic API key not found.\n"
                "Please add your key to anthropic_api.txt,\n"
                "or switch to Manual mode which requires no API key."
            )

    # 2. Read leads
    log("📖 Reading Time_Vehicle_Leads.xlsx...")
    leads, subject, body_template = read_leads(
        base_dir, email_selection, progress_callback
    )

    if not leads:
        raise ValueError(
            "No matching emails found.\n"
            "Check that the email addresses exist in the leads file."
        )

    mode_label = "AI personalisation" if draft_mode == "AI" else "Manual template"
    log(f"✉️  Mode: {mode_label}")
    log("")

    # 3. Gmail auth
    log("🔐 Connecting to Gmail...")
    service = authenticate_gmail(base_dir, progress_callback)
    log("")
    log(f"🚀 Creating {len(leads)} drafts...")
    log("─" * 50)

    # 4. Draft loop
    created = 0
    failed  = 0

    for i, lead in enumerate(leads, 1):
        name  = lead["name"]
        email = lead["email"]
        try:
            if draft_mode == "AI":
                log(f"✍️  [{i}/{len(leads)}] AI personalising for: {name}")
                final_body = rewrite_body(api_key, body_template, lead)
            else:
                log(f"✍️  [{i}/{len(leads)}] Preparing draft for: {name}")
                final_body = apply_manual_template(body_template, lead)

            create_draft(service, email, subject, final_body)
            created += 1
            log(f"   ✅ Draft → {email}")
            time.sleep(0.3)
        except Exception as e:
            failed += 1
            log(f"   ❌ Failed ({email}): {str(e)[:80]}")

    log("")
    log(f"🎯 DONE — {created} drafts created" +
        (f", {failed} failed" if failed else " — all successful!"))

    return {"created": created, "failed": failed, "total": len(leads)}
