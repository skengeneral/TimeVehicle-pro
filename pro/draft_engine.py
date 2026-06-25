import os
import sys
import base64
import time
import requests
from pathlib import Path
from email.mime.text import MIMEText
from openpyxl import load_workbook

# Gmail API
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.compose']

# ── Path helpers ──────────────────────────────────────────────────
def _base(base_dir=None):
    if base_dir:
        return Path(base_dir)
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(os.getcwd())

def get_anthropic_key(base_dir=None):
    f = _base(base_dir) / "anthropic_api.txt"
    if f.exists():
        return f.read_text(encoding="utf-8").strip()
    return os.environ.get("ANTHROPIC_API_KEY", "")

def get_credentials_path(base_dir=None):
    return _base(base_dir) / "credentials.json"

def get_token_path(base_dir=None):
    return _base(base_dir) / "gmail_token.json"

# ── Gmail auth ────────────────────────────────────────────────────
def authenticate_gmail(base_dir=None, progress_callback=None):
    def log(m):
        if progress_callback: progress_callback(m)

    creds_path = get_credentials_path(base_dir)
    token_path = get_token_path(base_dir)

    if not creds_path.exists():
        raise FileNotFoundError(
            "credentials.json not found.\n"
            "Please contact Time Vehicle support to obtain this file."
        )

    creds = None
    if token_path.exists():
        try:
            creds = Credentials.from_authorized_user_file(str(token_path), GMAIL_SCOPES)
        except Exception:
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            log("🔄 Refreshing Gmail session...")
            creds.refresh(Request())
        else:
            log("🌐 Opening Gmail sign-in in your browser...")
            flow = InstalledAppFlow.from_client_secrets_file(
                str(creds_path), GMAIL_SCOPES
            )
            creds = flow.run_local_server(port=0, open_browser=True)
            log("✅ Gmail authorization granted")

        token_path.write_text(creds.to_json())

    return build("gmail", "v1", credentials=creds)

def get_gmail_profile(base_dir=None):
    """Returns the connected Gmail address, or None if not yet authorized."""
    token_path = get_token_path(base_dir)
    creds_path = get_credentials_path(base_dir)
    if not token_path.exists() or not creds_path.exists():
        return None
    try:
        creds = Credentials.from_authorized_user_file(str(token_path), GMAIL_SCOPES)
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        if creds and creds.valid:
            svc     = build("gmail", "v1", credentials=creds)
            profile = svc.users().getProfile(userId="me").execute()
            return profile.get("emailAddress")
    except Exception:
        pass
    return None

# ── Claude API body rewriter ──────────────────────────────────────
def rewrite_body(api_key, body_template, business_name):
    """
    Calls Claude claude-sonnet-4-6 to produce a unique, personalised version
    of the client's body template for each business/professional.
    """
    prompt = (
        f"You are helping a professional send personalised outreach emails.\n\n"
        f"The recipient is: {business_name}\n\n"
        f"Original body template:\n---\n{body_template}\n---\n\n"
        f"Rewrite the body so it is unique and personalised for {business_name}:\n"
        f"- Naturally mention '{business_name}' once or twice where it fits organically\n"
        f"- Vary sentence structure, word choices, and phrasing from the original\n"
        f"- Keep the exact same core message, intent, tone, and approximate length\n"
        f"- Sound professional and human — not AI-generated\n"
        f"- Do NOT add a greeting (Dear...) or sign-off — body paragraphs only\n"
        f"- Do NOT add a subject line\n"
        f"- Return ONLY the rewritten body text, nothing else"
    )

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "Content-Type": "application/json",
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01"
        },
        json={
            "model": "claude-sonnet-4-6",
            "max_tokens": 1000,
            "messages": [{"role": "user", "content": prompt}]
        },
        timeout=30
    )
    if resp.status_code == 200:
        return resp.json()["content"][0]["text"].strip()
    raise Exception(f"Claude API error {resp.status_code}: {resp.text[:200]}")

# ── Gmail draft creator ───────────────────────────────────────────
def create_draft(service, to_email, subject, body):
    msg             = MIMEText(body, "plain", "utf-8")
    msg["to"]       = to_email
    msg["subject"]  = subject
    raw             = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    draft           = service.users().drafts().create(
        userId="me", body={"message": {"raw": raw}}
    ).execute()
    return draft.get("id")

# ── Leads reader ──────────────────────────────────────────────────
def read_leads(base_dir=None, progress_callback=None):
    def log(m):
        if progress_callback: progress_callback(m)

    leads_file = _base(base_dir) / "Time_Vehicle_Leads.xlsx"
    if not leads_file.exists():
        raise FileNotFoundError(
            "Time_Vehicle_Leads.xlsx not found.\n"
            "Please run a search first to generate the leads file."
        )

    wb   = load_workbook(str(leads_file))
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise ValueError("Leads file appears empty. Please run a new search.")

    # Pull subject and body from the bottom section
    subject       = ""
    body_template = ""
    for row in rows:
        if not row or not row[0]:
            continue
        label = str(row[0]).strip().upper()
        if "MAIL SUBJECT" in label and len(row) > 1:
            subject = str(row[1] or "").strip()
        elif "MAIL BODY" in label and len(row) > 1:
            body_template = str(row[1] or "").strip()

    if not subject:
        raise ValueError(
            "MAIL SUBJECT is empty.\n"
            "Please fill it in Time_Vehicle_Leads.xlsx and save."
        )
    if not body_template:
        raise ValueError(
            "MAIL BODY TEMPLATE is empty.\n"
            "Please fill it in Time_Vehicle_Leads.xlsx and save."
        )

    # Pull selected rows
    header_row     = rows[0]
    selected_leads = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        if str(row[0]).strip().upper() != "YES":
            continue
        lead  = {str(h): v for h, v in zip(header_row, row) if h}
        email = str(lead.get("Email ID", "") or "").strip()
        name  = str(lead.get("Business Name", "") or "").strip()
        if email and email.lower() != "not provided" and "@" in email:
            selected_leads.append({"name": name, "email": email})

    log(f"📋 {len(selected_leads)} emails selected for drafting")
    log(f"📝 Subject: {subject}")
    return selected_leads, subject, body_template

# ── Main entry point ──────────────────────────────────────────────
def create_bulk_drafts(base_dir=None, progress_callback=None):
    """
    Called from the UI Tab 2 worker thread.
    Reads Time_Vehicle_Leads.xlsx, rewrites each body with Claude,
    and creates Gmail drafts for every selected email.
    """
    def log(m):
        if progress_callback: progress_callback(m)

    # 1. API key check
    api_key = get_anthropic_key(base_dir)
    if not api_key:
        raise ValueError(
            "Anthropic API key not found.\n"
            "Please add your key to anthropic_api.txt and try again."
        )

    # 2. Read leads + selections
    log("📖 Reading Time_Vehicle_Leads.xlsx...")
    leads, subject, body_template = read_leads(base_dir, progress_callback)

    if not leads:
        raise ValueError(
            "No emails selected.\n"
            "Please set SELECT = Yes on at least one row in Time_Vehicle_Leads.xlsx."
        )

    log("")

    # 3. Gmail auth
    log("🔐 Connecting to Gmail...")
    service = authenticate_gmail(base_dir, progress_callback)
    log("")
    log(f"🚀 Creating {len(leads)} personalised drafts...")
    log("─" * 50)

    # 4. Draft creation loop
    created = 0
    failed  = 0

    for i, lead in enumerate(leads, 1):
        name  = lead["name"]
        email = lead["email"]
        try:
            log(f"✍️  [{i}/{len(leads)}] Personalising for: {name}")
            unique_body = rewrite_body(api_key, body_template, name)
            create_draft(service, email, subject, unique_body)
            created += 1
            log(f"   ✅ Draft → {email}")
            time.sleep(0.3)          # gentle rate limiting
        except Exception as e:
            failed += 1
            log(f"   ❌ Failed ({email}): {str(e)[:80]}")

    log("")
    log(f"🎯 DONE — {created} drafts created" +
        (f", {failed} failed" if failed else " — all successful!"))

    return {"created": created, "failed": failed, "total": len(leads)}
